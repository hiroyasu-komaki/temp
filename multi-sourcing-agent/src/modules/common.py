# -*- coding: utf-8 -*-
"""
vmo_impact_model.xlsx の入力値を読み込み、「計算」シート・「投資評価」シートと
同じ数式をPythonで再現するための共通モジュール。

xlsxのセル(前提シートの青字/黄色セル)を書き換えて保存すれば、
このモジュールを経由する4つのグラフ生成スクリプトが自動的に最新の数値で
PNGを再作成する。

依存パッケージ: openpyxl, matplotlib, numpy
    pip install openpyxl matplotlib numpy --break-system-packages
"""
from __future__ import annotations

import json
import os
from dataclasses import asdict, dataclass, field

import matplotlib
import numpy as np
import openpyxl

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402

# ---------------------------------------------------------------------------
# パス関連
# ---------------------------------------------------------------------------
MODULES_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(os.path.dirname(MODULES_DIR))
DEFAULT_XLSX_PATH = os.path.join(PROJECT_DIR, "input", "vmo_impact_model.xlsx")
DEFAULT_MID_DIR = os.path.join(PROJECT_DIR, "mid")
DEFAULT_OUTPUT_DIR = os.path.join(PROJECT_DIR, "output")

# 中間ファイル(mid/)のファイル名。各グラフ生成スクリプトはxlsxを直接読まず、
# これらのJSONだけを参照する(xlsx→JSONの変換内容を目視で検証できるようにするため)。
INPUTS_JSON = "inputs.json"
CALC_JSON = "calc.json"
INVESTMENT_EVAL_JSON = "investment_eval.json"

# 中間ファイルに書き出すgross_reduction_by_yearの年数。
# ②の仮説チャート(Y1-Y5表示)が必要とする5年分をここで確定させる。
MID_EXTRA_YEARS = 5

# ---------------------------------------------------------------------------
# 日本語フォント設定 (Mac/Win/Linuxで使えるフォントを順に探す)
# ---------------------------------------------------------------------------
_JP_FONT_CANDIDATES = [
    "Hiragino Sans",
    "Hiragino Kaku Gothic Pro",
    "Yu Gothic",
    "Meiryo",
    "Noto Sans CJK JP",
    "IPAexGothic",
    "MS Gothic",
    "TakaoGothic",
]


def setup_japanese_font() -> None:
    """利用可能な日本語フォントを rcParams に設定する。"""
    from matplotlib import font_manager

    available = {f.name for f in font_manager.fontManager.ttflist}
    for name in _JP_FONT_CANDIDATES:
        if name in available:
            plt.rcParams["font.family"] = name
            break
    else:
        # 見つからない場合でもフォールバックで文字化けよりは欠字の方がまし
        plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["figure.dpi"] = 150
    plt.rcParams["savefig.dpi"] = 150


# ---------------------------------------------------------------------------
# カテゴリの表示順・配色 (前提シート B21:B26 の並び順に合わせる)
# ---------------------------------------------------------------------------
CATEGORY_ORDER = [
    "既存システム保守運用費",
    "システム構築・導入費",
    "ソフトウェアライセンス費",
    "クラウド利用料",
    "ネットワーク機器・回線",
    "EUC(PC/スマホ)レンタル",
]

CATEGORY_COLORS = {
    "既存システム保守運用費": "#1f4e6b",
    "システム構築・導入費": "#2e7ba6",
    "ソフトウェアライセンス費": "#5fa8d3",
    "クラウド利用料": "#a8d4e8",
    "ネットワーク機器・回線": "#a9d6b0",
    "EUC(PC/スマホ)レンタル": "#c9a227",
}

# 「①事実提示」チャートの契約更新到来時期(Y1/Y2/Y3/Y4以降)への配分比率。
# 前提シートには契約単位の更新月データが無いため、平均契約年数だけでは
# 一意に決まらない。元の資料では「一般的代表値のプレースホルダ(実データ
# 差込で確定)」と明記されている配分を採用し、全カテゴリに同一比率で適用する。
RENEWAL_TIMING_RATIOS = {
    "Y1": 2858 / 6400,
    "Y2": 1805 / 6400,
    "Y3": 1158 / 6400,
    "Y4以降": 579 / 6400,
}


# ---------------------------------------------------------------------------
# xlsx読み込み
# ---------------------------------------------------------------------------
@dataclass
class CategoryInput:
    name: str
    spend_ratio: float
    competitive_ratio: float
    reduction_conservative: float
    reduction_mid: float
    reduction_optimistic: float
    avg_contract_years: float


@dataclass
class ModelInputs:
    it_spend: float
    sole_source_ratio: float
    wacc: float
    eval_years: float
    prep_cost: float
    vmo_initial_cost: float
    vmo_opex: float
    rfp_cost: list  # [Y1, Y2, Y3]
    categories: list = field(default_factory=list)


def load_inputs(xlsx_path: str = DEFAULT_XLSX_PATH) -> ModelInputs:
    """「前提」シートの生数値(青字/黄色セル)を読み込む。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["前提"]

    categories = []
    for row in range(21, 27):
        categories.append(
            CategoryInput(
                name=ws[f"B{row}"].value,
                spend_ratio=ws[f"C{row}"].value,
                competitive_ratio=ws[f"D{row}"].value,
                reduction_conservative=ws[f"E{row}"].value,
                reduction_mid=ws[f"F{row}"].value,
                reduction_optimistic=ws[f"G{row}"].value,
                avg_contract_years=ws[f"H{row}"].value,
            )
        )

    return ModelInputs(
        it_spend=ws["C5"].value,
        sole_source_ratio=ws["C6"].value,
        wacc=ws["C7"].value,
        eval_years=ws["C8"].value,
        prep_cost=ws["C11"].value,
        vmo_initial_cost=ws["C12"].value,
        vmo_opex=ws["C13"].value,
        rfp_cost=[ws["C14"].value, ws["C15"].value, ws["C16"].value],
        categories=categories,
    )


# ---------------------------------------------------------------------------
# 「計算」シートの再現
# ---------------------------------------------------------------------------
@dataclass
class CategoryCalc:
    name: str
    spend: float  # 支出額
    sole_source: float  # 随意契約額(事実)
    competitive: float  # 競争対象額(合意割合)
    reduction_conservative: float  # 年間削減:保守 (フル・ランプ前)
    reduction_mid: float  # 年間削減:中位 (フル・ランプ前)
    reduction_optimistic: float  # 年間削減:楽観 (フル・ランプ前)
    gross_reduction_by_year: dict  # {1: Y1グロス削減(中位,ランプ後), 2: ..., 3: ...}


def compute_calc_sheet(inputs: ModelInputs, extra_years: int = 3) -> list:
    """「計算」シートのB列〜J列を数式どおりに再現する。

    extra_years: 何年目までグロス削減のランプアップを計算するか
                 (xlsxの評価期間は3年だが、②の仮説チャートは5年分表示するため
                 min(t/平均契約年数, 1) の式をそのまま延長して使う)
    """
    rows = []
    for cat in inputs.categories:
        spend = inputs.it_spend * cat.spend_ratio
        sole_source = spend * inputs.sole_source_ratio
        competitive = sole_source * cat.competitive_ratio
        reduction_conservative = competitive * cat.reduction_conservative
        reduction_mid = competitive * cat.reduction_mid
        reduction_optimistic = competitive * cat.reduction_optimistic

        gross_by_year = {}
        for t in range(1, extra_years + 1):
            ramp = min(t / cat.avg_contract_years, 1.0)
            gross_by_year[t] = reduction_mid * ramp

        rows.append(
            CategoryCalc(
                name=cat.name,
                spend=spend,
                sole_source=sole_source,
                competitive=competitive,
                reduction_conservative=reduction_conservative,
                reduction_mid=reduction_mid,
                reduction_optimistic=reduction_optimistic,
                gross_reduction_by_year=gross_by_year,
            )
        )
    return rows


# ---------------------------------------------------------------------------
# 「投資評価」シートの再現 (NPV / IRR / PI / 割引回収期間)
# ---------------------------------------------------------------------------
@dataclass
class InvestmentEval:
    gross_reduction: dict  # {1,2,3: グロス削減}
    rfp_cost: dict  # {1,2,3: RFP実行コスト}
    vmo_opex: dict  # {1,2,3: VMO運営費}
    net_cf: dict  # {0,1,2,3: ネットCF}
    discount_factor: dict
    discounted_cf: dict
    cumulative_npv: dict  # 割引後累積
    cumulative_cf_undiscounted: dict  # 割引前累積
    npv: float
    irr: float
    pi: float
    payback_years: float  # 割引回収期間(線形補間)


def _irr_bisection(cashflows: list, lo: float = -0.9, hi: float = 5.0, tol: float = 1e-8) -> float:
    """cashflows[0]がY0(通常マイナス)から始まる年次CFのIRRをNPV=0となる割引率として求める。"""

    def npv_at(rate):
        return sum(cf / (1 + rate) ** t for t, cf in enumerate(cashflows))

    f_lo, f_hi = npv_at(lo), npv_at(hi)
    if f_lo * f_hi > 0:
        # 探索範囲を広げる
        hi = 50.0
        f_hi = npv_at(hi)
        if f_lo * f_hi > 0:
            return float("nan")
    for _ in range(200):
        mid = (lo + hi) / 2
        f_mid = npv_at(mid)
        if abs(f_mid) < tol:
            return mid
        if f_lo * f_mid < 0:
            hi, f_hi = mid, f_mid
        else:
            lo, f_lo = mid, f_mid
    return (lo + hi) / 2


def compute_investment_eval(inputs: ModelInputs, calc_rows: list) -> InvestmentEval:
    gross_total = {t: sum(r.gross_reduction_by_year[t] for r in calc_rows) for t in (1, 2, 3)}
    rfp_cost = {1: inputs.rfp_cost[0], 2: inputs.rfp_cost[1], 3: inputs.rfp_cost[2]}
    vmo_opex = {t: inputs.vmo_opex for t in (1, 2, 3)}

    net_cf = {0: -(inputs.prep_cost + inputs.vmo_initial_cost)}
    for t in (1, 2, 3):
        net_cf[t] = gross_total[t] - rfp_cost[t] - vmo_opex[t]

    wacc = inputs.wacc
    discount_factor = {t: 1 / (1 + wacc) ** t for t in (0, 1, 2, 3)}
    discounted_cf = {t: net_cf[t] * discount_factor[t] for t in (0, 1, 2, 3)}

    cumulative_npv = {}
    running = 0.0
    for t in (0, 1, 2, 3):
        running += discounted_cf[t]
        cumulative_npv[t] = running

    cumulative_cf_undiscounted = {}
    running_u = 0.0
    for t in (0, 1, 2, 3):
        running_u += net_cf[t]
        cumulative_cf_undiscounted[t] = running_u

    npv = cumulative_npv[3]
    irr = _irr_bisection([net_cf[0], net_cf[1], net_cf[2], net_cf[3]])
    pi = (cumulative_npv[3] + abs(net_cf[0])) / abs(net_cf[0])

    # 割引回収期間: 累積NPVが0を超える年を線形補間で求める
    payback_years = float("nan")
    prev_t, prev_val = 0, cumulative_npv[0]
    for t in (1, 2, 3):
        cur_val = cumulative_npv[t]
        if prev_val <= 0 < cur_val:
            payback_years = prev_t + (0 - prev_val) / (cur_val - prev_val)
            break
        prev_t, prev_val = t, cur_val

    return InvestmentEval(
        gross_reduction=gross_total,
        rfp_cost=rfp_cost,
        vmo_opex=vmo_opex,
        net_cf=net_cf,
        discount_factor=discount_factor,
        discounted_cf=discounted_cf,
        cumulative_npv=cumulative_npv,
        cumulative_cf_undiscounted=cumulative_cf_undiscounted,
        npv=npv,
        irr=irr,
        pi=pi,
        payback_years=payback_years,
    )


def ensure_output_dir(output_dir: str = DEFAULT_OUTPUT_DIR) -> str:
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


# ---------------------------------------------------------------------------
# 中間ファイル(mid/*.json)の書き出し・読み込み
#
# xlsx読み込みと数式の再現(compute_calc_sheet / compute_investment_eval)は
# ここで一度だけ実行し、結果をJSONとしてmid/に固定する。
# 各グラフ生成スクリプト(chart0X)はxlsxを直接読まず、このJSONだけを参照する。
# こうすることで、xlsxからの転記結果を後から目視・diffで検証できる。
# ---------------------------------------------------------------------------
def ensure_mid_dir(mid_dir: str = DEFAULT_MID_DIR) -> str:
    os.makedirs(mid_dir, exist_ok=True)
    return mid_dir


def _write_json(path: str, data) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _read_json(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _with_int_keys(d: dict) -> dict:
    """JSON読み込み後、年次キー("0","1",...)をintに戻す。"""
    return {int(k): v for k, v in d.items()}


def build_mid_files(
    xlsx_path: str = DEFAULT_XLSX_PATH,
    mid_dir: str = DEFAULT_MID_DIR,
    extra_years: int = MID_EXTRA_YEARS,
) -> dict:
    """xlsxを読み込み、計算結果を中間JSON(mid/)に書き出す。

    書き出すファイル:
        inputs.json           前提シートの生数値(ModelInputs)
        calc.json              計算シート相当の再現結果(CategoryCalcのリスト)
        investment_eval.json  投資評価シート相当の再現結果(InvestmentEval)
    """
    mid_dir = ensure_mid_dir(mid_dir)

    inputs = load_inputs(xlsx_path)
    calc_rows = compute_calc_sheet(inputs, extra_years=extra_years)
    inv_eval = compute_investment_eval(inputs, calc_rows)

    inputs_path = os.path.join(mid_dir, INPUTS_JSON)
    calc_path = os.path.join(mid_dir, CALC_JSON)
    inv_eval_path = os.path.join(mid_dir, INVESTMENT_EVAL_JSON)

    _write_json(inputs_path, asdict(inputs))
    _write_json(calc_path, [asdict(r) for r in calc_rows])
    _write_json(inv_eval_path, asdict(inv_eval))

    return {
        "inputs": inputs_path,
        "calc": calc_path,
        "investment_eval": inv_eval_path,
    }


def _require_mid_file(path: str) -> str:
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"中間ファイルが見つかりません: {path}\n"
            "先に build_mid.py (または generate_all.py) を実行してxlsxから中間JSONを作成してください。"
        )
    return path


def load_inputs_from_mid(mid_dir: str = DEFAULT_MID_DIR) -> ModelInputs:
    """mid/inputs.json からModelInputsを復元する。"""
    path = _require_mid_file(os.path.join(mid_dir, INPUTS_JSON))
    data = dict(_read_json(path))
    data["categories"] = [CategoryInput(**c) for c in data["categories"]]
    return ModelInputs(**data)


def load_calc_from_mid(mid_dir: str = DEFAULT_MID_DIR) -> list:
    """mid/calc.json からCategoryCalcのリストを復元する。"""
    path = _require_mid_file(os.path.join(mid_dir, CALC_JSON))
    rows = []
    for r in _read_json(path):
        r = dict(r)
        r["gross_reduction_by_year"] = _with_int_keys(r["gross_reduction_by_year"])
        rows.append(CategoryCalc(**r))
    return rows


def load_investment_eval_from_mid(mid_dir: str = DEFAULT_MID_DIR) -> InvestmentEval:
    """mid/investment_eval.json からInvestmentEvalを復元する。"""
    path = _require_mid_file(os.path.join(mid_dir, INVESTMENT_EVAL_JSON))
    data = dict(_read_json(path))
    for key in (
        "gross_reduction",
        "rfp_cost",
        "vmo_opex",
        "net_cf",
        "discount_factor",
        "discounted_cf",
        "cumulative_npv",
        "cumulative_cf_undiscounted",
    ):
        data[key] = _with_int_keys(data[key])
    return InvestmentEval(**data)
